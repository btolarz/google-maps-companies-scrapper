import time
import re
import requests
import argparse
import logging
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'accept-language': 'pl-PL,pl;q=0.9'
}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

sheet_name = 'data.xlsx'

# Write Headline and create a new excel sheet
def xl_sheet_headlines(sheet_name=sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = ['url', 'name', 'address', 'website', 'phone', 'email']
    ws.append(headlines)
    wb.save(sheet_name)
    logging.info("Created new Excel sheet with headlines.")

xl_sheet_headlines()

# Write Data On existing sheet
def xl_write(data_write, sheet_name=sheet_name):
    wb = load_workbook(sheet_name)
    work_sheet = wb.active  # Get active sheet
    work_sheet.append(data_write)
    wb.save(sheet_name)

def driver_define():
    logging.info('Firefox Browser Opening')
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--lang=pl-PL')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36')
    options.binary_location = r'/usr/bin/firefox-esr'
    service = Service('/usr/bin/geckodriver')
    driver = webdriver.Firefox(service=service, options=options)
    return driver

# Email and Phone Get from Website
def get_email_and_phone_from_website(url):
    try:
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.text, 'lxml')
        email = None
        phone = None
        # Search for mailto links
        for mailto in soup.select('a[href^=mailto]'):
            email = mailto['href'].split(':')[1].split('?')[0]  # Split to remove any parameters
            if email:
                break
        # If no mailto links found, search for email and phone patterns in text
        if not email:
            emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", soup.text)
            if emails:
                email = emails[0]
        if not phone:
            phones = re.findall(r"\+?\d[\d\s-]{8,}\d", soup.text)
            if phones:
                phone = phones[0]
        # Search for contact links
        if not email or not phone:
            contact_link = None
            for link in soup.find_all('a', href=True):
                if any(keyword in link.text.lower() for keyword in ['kontakt', 'contact', 'CONTACT']):
                    contact_link = link['href']
                    if not contact_link.startswith('http'):
                        contact_link = requests.compat.urljoin(url, contact_link)
                    break
            if contact_link:
                contact_response = requests.get(contact_link, headers=headers)
                contact_soup = BeautifulSoup(contact_response.text, 'lxml')
                if not email:
                    emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", contact_soup.text)
                    if emails:
                        email = emails[0]
                if not phone:
                    phones = re.findall(r"\+?\d[\d\s-]{8,}\d", contact_soup.text)
                    if phones:
                        phone = phones[0]
        return email if email else "", phone if phone else ""
    except Exception as e:
        logging.error(f"Error fetching email or phone from website {url}: {e}")
        return "", ""

# Email and Phone Get from Facebook
def get_email_and_phone_from_facebook(driver, url):
    try:
        driver.get(url)
        time.sleep(3)
        page_source = driver.page_source
        emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", page_source)
        phones = re.findall(r"\+?\d[\d\s-]{8,}\d", page_source)
        email = emails[0] if emails else ""
        phone = phones[0] if phones else ""
        return email, phone
    except Exception as e:
        logging.error(f"Error fetching email or phone from Facebook page {url}: {e}")
        return "", ""

def clean_text(text):
    # Clean unwanted characters
    return text.replace(u'\u2022', '').replace(u'\u2192', '').replace('\ue0c8\n', '').replace('\ue0b0\n', '').strip()

# Main function
def main(start_url):
    driver = driver_define()
    found_urls = []

    driver.get(start_url)
    time.sleep(1)

    try:
        driver.find_element(By.XPATH, '//span[text()="Zaakceptuj wszystko"]').click()
    except:
        pass

    time.sleep(1)

    container = driver.find_element(By.XPATH, '//div[text()="Wyznacz trasę"]/parent::*/parent::*/parent::*/parent::*/parent::*/parent::*/parent::*/parent::*')
    last_height = driver.execute_script("return arguments[0].scrollHeight", container)
    while True:
        logging.info("Scrolling to load all results...")
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", container)
        time.sleep(2)
        new_height = driver.execute_script("return arguments[0].scrollHeight", container)
        if new_height == last_height:
            break
        last_height = new_height

    logging.info("Scrolling completed. Fetching URLs...")
    elements = driver.find_elements(By.XPATH, '//div[text()="Wyznacz trasę"]/parent::*/parent::*/parent::*/parent::*/parent::*/parent::*')

    logging.info(f"Found {len(elements)} URLs.")

    for elem in elements:
        try:
            href = elem.find_element(By.XPATH, './/a').get_attribute('href')
            found_urls.append(href)
        except:
            continue

    logging.info("Fetching contact details...")
    for found_url in found_urls:
        try:
            logging.info(f"------- {found_urls.index(found_url)+1} of {len(found_urls)} -------")
            driver.get(found_url)
            time.sleep(1)
            
            try:
                name = clean_text(driver.find_element(By.CSS_SELECTOR, 'h1').text)
            except Exception as e:
                logging.error(f"Error fetching name from {found_url}: {e}")
                name = ''

            try:
                address = clean_text(driver.find_element(By.CSS_SELECTOR, '[data-tooltip="Kopiuj adres"]').text)
            except Exception as e:
                logging.error(f"Error fetching address from {found_url}: {e}")
                address = ''

            try:
                phone = clean_text(driver.find_element(By.CSS_SELECTOR, '[data-tooltip="Kopiuj numer telefonu"]').text)
            except Exception as e:
                phone = ''

            try:
                website = driver.find_element(By.CSS_SELECTOR, '[data-tooltip="Otwórz witrynę"]').get_attribute('href')
            except Exception as e:
                website = ''

            email = ''
            if website:
                if "facebook.com" in website:
                    email, phone_from_fb = get_email_and_phone_from_facebook(driver, website)
                    if not phone:
                        phone = phone_from_fb
                else:
                    email, phone_from_site = get_email_and_phone_from_website(website)
                    if not phone:
                        phone = phone_from_site

            logging.info(f"name: {name}")
            logging.info(f"address: {address}")
            logging.info(f"website: {website}")
            logging.info(f"phone: {phone}")
            logging.info(f"email: {email}")

            write_data = [found_url, name, address, website, phone, email]
            xl_write(write_data)

        except Exception as e:
            logging.error(f"Error processing URL {found_url}: {e}")
            continue

    driver.quit()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape a website for contact details.')
    parser.add_argument('start_url', type=str, help='The starting URL for the scraper')
    args = parser.parse_args()
    main(args.start_url)
